import os
import json
from datetime import datetime

import openpyxl
import requests
import yaml
from twilio.rest import Client
from twilio.base.exceptions import TwilioRestException


# ---------- Config ----------
def load_config(path: str = "config.yaml") -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Config file not found: {path}")

    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}

    # Twilio section with env var overrides
    twilio_cfg = cfg.get("twilio", {}) or {}
    twilio_cfg["account_sid"] = os.getenv("TWILIO_ACCOUNT_SID", twilio_cfg.get("account_sid"))
    twilio_cfg["auth_token"] = os.getenv("TWILIO_AUTH_TOKEN", twilio_cfg.get("auth_token"))
    twilio_cfg["from_number"] = os.getenv("TWILIO_FROM_NUMBER", twilio_cfg.get("from_number"))

    if not twilio_cfg.get("account_sid"):
        raise ValueError("Missing Twilio Account SID (set TWILIO_ACCOUNT_SID or 'twilio.account_sid' in config.yaml).")
    if not twilio_cfg.get("auth_token"):
        raise ValueError("Missing Twilio Auth Token (set TWILIO_AUTH_TOKEN or 'twilio.auth_token' in config.yaml).")
    if not twilio_cfg.get("from_number"):
        raise ValueError("Missing Twilio from_number (set TWILIO_FROM_NUMBER or 'twilio.from_number' in config.yaml).")

    cfg["twilio"] = twilio_cfg

    # WhatsApp (Meta) section with env var override
    wa_cfg = cfg.get("whatsapp", {}) or {}
    if not wa_cfg.get("api_url"):
        raise ValueError("Missing WhatsApp API URL ('whatsapp.api_url' in config.yaml).")
    wa_cfg["access_token"] = os.getenv("WHATSAPP_ACCESS_TOKEN", wa_cfg.get("access_token"))
    if not wa_cfg.get("access_token"):
        raise ValueError(
            "Missing WhatsApp access token (set WHATSAPP_ACCESS_TOKEN or 'whatsapp.access_token' in config.yaml)."
        )
    cfg["whatsapp"] = wa_cfg

    # Schedule / workbook
    sched = cfg.get("schedule", {}) or {}
    if not sched.get("workbook"):
        raise ValueError("Missing rota workbook path ('schedule.workbook' in config.yaml).")
    cfg["schedule"] = sched

    # Contacts
    cfg["contacts"] = cfg.get("contacts", {}) or {}

    return cfg


CONFIG = load_config()  # load once per container


# ---------- Helpers ----------
def get_mobile_number(name: str) -> str:
    return CONFIG["contacts"].get(name, "+44XXXXXXXXXX")


def build_twilio_client() -> Client:
    return Client(CONFIG["twilio"]["account_sid"], CONFIG["twilio"]["auth_token"])


def send_whatsapp_message_meta(phone_number: str, message: str) -> requests.Response:
    """Send WhatsApp via Meta Graph API (optional path if you want to use it)."""
    headers = {
        "Authorization": f"Bearer {CONFIG['whatsapp']['access_token']}",
        "Content-Type": "application/json",
    }
    data = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "preview_url": False,
        "to": phone_number,
        "type": "text",
        "text": {"body": message},
    }
    return requests.post(CONFIG["whatsapp"]["api_url"], headers=headers, json=data, timeout=15)


def send_whatsapp_message_twilio(twilio_client: Client, phone_number: str, message_body: str):
    """Send WhatsApp via Twilio (requires an approved WhatsApp sender)."""
    try:
        msg = twilio_client.messages.create(
            body=message_body,
            from_=f"whatsapp:{CONFIG['twilio']['from_number']}",
            to=f"whatsapp:{phone_number}",
        )
        return msg.sid
    except TwilioRestException as e:
        print(f"[ERROR] WhatsApp (Twilio) send failed: {e}")
        return None


def send_text_message_twilio(twilio_client: Client, phone_number: str, message_body: str):
    """Send SMS via Twilio."""
    try:
        msg = twilio_client.messages.create(
            body=message_body,
            from_=CONFIG["twilio"]["from_number"],
            to=phone_number,
        )
        return msg.sid
    except TwilioRestException as e:
        print(f"[ERROR] SMS (Twilio) send failed: {e}")
        return None


# ---------- Lambda handler ----------
def lambda_handler(event, context):
    # Load rota workbook
    try:
        workbook_path = CONFIG["schedule"]["workbook"]
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        sheet = workbook.active
    except Exception as e:
        print(f"[ERROR] Loading Excel file: {e}")
        return {"status": "error", "reason": "excel_load_failed"}

    today = datetime.now().date()

    # Init Twilio client (done here to ensure fresh credentials if rotated)
    twilio_client = build_twilio_client()

    tasks_found = False

    # Iterate rows (assumes first row is header: A1='Dates', B1..=task names)
    for row in sheet.iter_rows(min_row=2):
        cell = row[0].value
        task_date = cell.date() if hasattr(cell, "date") else cell  # handle datetime or date
        if task_date == today:
            tasks_found = True
            # For each task column (skip first column)
            for col_idx, cell in enumerate(row[1:], start=2):
                person = cell.value
                if not person:
                    continue

                task_name = sheet.cell(row=1, column=col_idx).value
                phone_number = get_mobile_number(str(person).strip())

                message = (
                    f"Reminder for {person}: Please complete your cleaning task - {task_name} "
                    f"on {today.strftime('%Y-%m-%d')}\n"
                )

                # Choose one path: Twilio SMS, Twilio WhatsApp, or Meta WhatsApp.
                # Example below: Twilio SMS
                sid = send_text_message_twilio(twilio_client, phone_number, message)

                # For Twilio WhatsApp instead, use:
                # sid = send_whatsapp_message_twilio(twilio_client, phone_number, message)

                # For Meta WhatsApp instead, use:
                # resp = send_whatsapp_message_meta(phone_number, message)
                # sid = "meta_ok" if resp.status_code == 200 else None

                if sid:
                    print(f"[OK] Sent to {person}: {sid}")
                else:
                    print(f"[FAIL] Could not send to {person}")

    if not tasks_found:
        print("No tasks found for today")
        return {"status": "ok", "messages_sent": 0}

    return {"status": "ok"}
    

# Local test (comment out on AWS if desired)
if __name__ == "__main__":
    print(json.dumps(lambda_handler(None, None)))
